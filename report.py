from openpyxl import Workbook


def save_report(trades, filename, result=None):
    """バックテスト結果をExcelに保存"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Backtest"

    ws.append([
        "No",
        "買い日",
        "売り日",
        "買値",
        "売値",
        "利益率(%)",
        "手数料込み利益率(%)",
        "手数料(円)",
        "決済後資産(円)",
        "保有日数",
        "決済理由"
    ])

    for i, trade in enumerate(trades, start=1):
        ws.append([
            i,
            str(trade["buy_date"].date()),
            str(trade["sell_date"].date()),
            round(trade["buy_price"], 2),
            round(trade["sell_price"], 2),
            round(trade.get("gross_profit", trade["profit"]), 2),
            round(trade["profit"], 2),
            round(trade.get("commission", 0), 0),
            round(trade.get("capital", 0), 0),
            trade["hold_days"],
            trade.get("exit_reason", "")
        ])

    ws["L1"] = "バックテスト結果"
    ws["L2"] = "総取引回数"
    ws["M2"] = len(trades)

    win_count = sum(1 for t in trades if t["profit"] > 0)
    win_rate = (win_count / len(trades) * 100) if trades else 0

    ws["L3"] = "勝率"
    ws["M3"] = f"{win_rate:.1f}%"

    total_profit = result["total_profit"] if result else sum(t["profit"] for t in trades)
    ws["L4"] = "総利益率"
    ws["M4"] = f"{total_profit:.2f}%"

    avg_hold = sum(t["hold_days"] for t in trades) / len(trades) if trades else 0
    ws["L5"] = "平均保有日数"
    ws["M5"] = f"{avg_hold:.1f}日"

    if result:
        ws["L6"] = "最終資産"
        ws["M6"] = round(result["final_capital"], 0)
        ws["L7"] = "手数料込み利益"
        ws["M7"] = round(result["net_profit_yen"], 0)
        ws["L8"] = "総手数料"
        ws["M8"] = round(result["total_commission"], 0)

    wb.save(filename)

    print(f"Excel保存完了: {filename}")
