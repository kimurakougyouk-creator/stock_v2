from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def format_signal_report(output_path: str | Path) -> None:
    """売買シグナルのExcelファイルを見やすく整える。"""
    path = Path(output_path)
    workbook = load_workbook(path)
    sheet = workbook.active
    sheet.title = "最新シグナル"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    buy_fill = PatternFill("solid", fgColor="C6EFCE")
    sell_fill = PatternFill("solid", fgColor="FFC7CE")
    hold_fill = PatternFill("solid", fgColor="E7E6E6")
    top_fill = PatternFill("solid", fgColor="FFF2CC")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 24

    headers = {
        cell.value: cell.column
        for cell in sheet[1]
        if cell.value is not None
    }

    signal_column = headers.get("Signal")
    rank_column = headers.get("Rank")

    for row in range(2, sheet.max_row + 1):
        signal = sheet.cell(row, signal_column).value if signal_column else None
        rank = sheet.cell(row, rank_column).value if rank_column else None

        if signal == "BUY":
            signal_fill = buy_fill
        elif signal == "SELL":
            signal_fill = sell_fill
        else:
            signal_fill = hold_fill

        if signal_column:
            sheet.cell(row, signal_column).fill = signal_fill
            sheet.cell(row, signal_column).font = Font(bold=True)
            sheet.cell(row, signal_column).alignment = Alignment(horizontal="center")

        if isinstance(rank, (int, float)) and rank <= 3:
            for cell in sheet[row]:
                if cell.column != signal_column:
                    cell.fill = top_fill

    currency_columns = {
        "Close",
        "MA5",
        "MA25",
        "MA75",
        "ATR",
        "StopPrice",
        "RiskPerShare",
        "MaxLossYen",
        "ReferenceAmountYen",
    }
    integer_columns = {"Score", "Rank", "ReferenceShares"}

    for header, column in headers.items():
        column_letter = get_column_letter(column)

        if header in currency_columns:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, column).number_format = '#,##0.0'
        elif header in integer_columns:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, column).number_format = '0'

        values = [
            str(sheet.cell(row, column).value or "")
            for row in range(1, sheet.max_row + 1)
        ]
        width = min(max(len(value) for value in values) + 2, 45)

        if header in {"Reason", "PositionSizingReason"}:
            width = 45
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, column).alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )

        sheet.column_dimensions[column_letter].width = max(width, 10)

    workbook.save(path)
